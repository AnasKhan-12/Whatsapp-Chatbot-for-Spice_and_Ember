"""
01_loaders.py — Load & Parse Data from PDF + Excel ONLY

REAL-WORLD SCENARIO:
  Your client sends you TWO files:
    - spice_and_ember_data.pdf   (5 pages, mixed formats)
    - spice_and_ember_menu.xlsx  (3 sheets, tabular data)

  Your job: extract EVERYTHING from these two files
  and turn each piece into a clean document with metadata.

PDF STRUCTURE (verified by inspection):
  Page 1 → About text (plain) + Key-Value table (contact/ops)
  Page 2 → Key-Value continued + Opening Hours table
  Page 3 → Full Menu table + FAQ (Q&A text)
  Page 4 → Chef's Notes (unstructured) + Nutrition header
  Page 5 → Nutrition inline data (pipe-separated text)

EXCEL STRUCTURE (verified by inspection):
  Sheet "Menu"           → 14 items, 10 columns
  Sheet "Nutrition"      → 12 items, 9 columns
  Sheet "Hours&Booking"  → 7 days, 5 columns

KEY LESSON:
  Before writing ANY loader, always inspect your files first.
  Never assume structure — always verify with code.
  (That's exactly what we did above with the inspection scripts.)
"""

import re
import pdfplumber
import openpyxl


# ══════════════════════════════════════════════════════════════
#  PDF LOADERS
#  We handle each page differently because each has
#  a different format — that's the whole point of this project.
# ══════════════════════════════════════════════════════════════

def load_pdf(pdf_path: str) -> list[dict]:
    """
    Master PDF loader — routes each page to the right sub-loader.
    Returns a flat list of documents, each being:
    {
        "page_content": str,   ← the text that will be embedded
        "metadata": dict       ← source info, type, page number, etc.
    }
    """
    all_docs = []

    with pdfplumber.open(pdf_path) as pdf:
        all_docs += _load_page1(pdf.pages[0])   # About + Contact KV table
        all_docs += _load_page2(pdf.pages[1])   # KV continued + Hours table
        all_docs += _load_page3(pdf.pages[2])   # Menu table + FAQ Q&A
        all_docs += _load_page4(pdf.pages[3])   # Chef's Notes
        all_docs += _load_page5(pdf.pages[4])   # Nutrition inline text

    print(f"✅ PDF: extracted {len(all_docs)} documents from {pdf_path}")
    return all_docs


# ── PAGE 1: Plain text (About) + Key-Value table (Contact) ────────────────────
def _load_page1(page) -> list[dict]:
    docs = []
    text = page.extract_text() or ""

    about_match = re.search(
        r"Section 1 — About the Restaurant.*?\n(.+?)(?=Section 2)",
        text, re.DOTALL
    )
    if about_match:
        about_text = about_match.group(1).strip()

        # BEFORE: splitting on \n broke mid-sentence
        # AFTER: join all lines into one string first
        # then split on ". " to get complete sentences
        one_string = " ".join(about_text.split("\n"))

        # Now split into real sentences using ". " as boundary
        # This keeps each sentence complete and meaningful
        sentences = [s.strip() + "." for s in one_string.split(". ") if len(s.strip()) > 30]

        for i, sentence in enumerate(sentences):
            docs.append({
                "page_content": sentence,
                "metadata": {
                    "source": "pdf",
                    "page": 1,
                    "section": "about_restaurant",
                    "format": "plain_text",
                    "doc_type": "restaurant_info",
                    "paragraph_index": i
                }
            })

    # ── Extract Key-Value table (Contact & Operations) ──
    tables = page.extract_tables()
    if tables:
        kv_table = tables[0]
        kv_lines = []
        for row in kv_table[1:]:   # skip header row ['Field', 'Value']
            if row[0] and row[1]:
                kv_lines.append(f"{row[0].strip()}: {row[1].strip()}")

        docs.append({
            "page_content": "Restaurant contact and operations information:\n" + "\n".join(kv_lines),
            "metadata": {
                "source": "pdf",
                "page": 1,
                "section": "contact_operations",
                "format": "key_value_table",
                "doc_type": "restaurant_info"
            }
        })

    return docs


# ── PAGE 2: KV table continued + Opening Hours table ──────────────────────────
def _load_page2(page) -> list[dict]:
    docs = []
    tables = page.extract_tables()

    # tables[0] = KV continuation (Founded, Head Chef, Mixologist)
    # tables[1] = Opening Hours (Day | Status | Open | Close)

    if len(tables) >= 1:
        kv_lines = []
        for row in tables[0]:
            if row and len(row) >= 2 and row[0] and row[1]:
                kv_lines.append(f"{str(row[0]).strip()}: {str(row[1]).strip()}")
        if kv_lines:
            docs.append({
                "page_content": "Additional restaurant info:\n" + "\n".join(kv_lines),
                "metadata": {
                    "source": "pdf",
                    "page": 2,
                    "section": "contact_operations_continued",
                    "format": "key_value_table",
                    "doc_type": "restaurant_info"
                }
            })

    # ── Opening Hours — one document per day for precise retrieval ──
    if len(tables) >= 2:
        hours_table = tables[1]
        for row in hours_table[1:]:     # skip header row
            if not row or not row[0]:
                continue
            day    = str(row[0]).strip()
            status = str(row[1]).strip()
            open_t = str(row[2]).strip()
            close_t = str(row[3]).strip()

            if status.upper() == "CLOSED":
                text = f"Opening hours: {day} — We are CLOSED. No delivery or dine-in."
            else:
                text = f"Opening hours: {day} — Open from {open_t} to {close_t}."

            docs.append({
                "page_content": text,
                "metadata": {
                    "source": "pdf",
                    "page": 2,
                    "section": "opening_hours",
                    "format": "table_row",
                    "doc_type": "hours",
                    "day": day,
                    "status": status
                }
            })

    return docs


# ── PAGE 3: Menu table + FAQ Q&A ──────────────────────────────────────────────
def _load_page3(page) -> list[dict]:
    docs = []
    text = page.extract_text() or ""

    # ── Menu table ──
    tables = page.extract_tables()
    if tables:
        menu_table = tables[0]
        for row in menu_table[1:]:     # skip header row
            if not row or not row[0] or str(row[0]).strip() == "ID":
                continue

            item_id  = str(row[0]).strip()
            name     = str(row[1]).strip()
            category = str(row[2]).strip()
            price    = str(row[3]).strip()
            cal      = str(row[4]).strip()
            spice    = str(row[5]).strip()
            veg      = str(row[6]).strip()
            gf       = str(row[7]).strip()

            # Convert table row → natural language
            # A query "do you have spicy food?" won't match "Hot" in a raw CSV
            # But it WILL match "spice level: Hot" in natural language
            text_item = (
                f"Menu item: {name} (ID: {item_id})\n"
                f"Category: {category}\n"
                f"Price: {price}\n"
                f"Calories: {cal}\n"
                f"Spice level: {spice}\n"
                f"Vegetarian: {veg} | Gluten free: {gf}"
            )

            docs.append({
                "page_content": text_item,
                "metadata": {
                    "source": "pdf",
                    "page": 3,
                    "section": "menu",
                    "format": "table_row",
                    "doc_type": "menu_item",
                    "item_id": item_id,
                    "item_name": name,
                    "category": category.lower(),
                    "price": price,
                    "spice_level": spice.lower(),
                    "is_vegetarian": veg,
                    "is_gluten_free": gf
                }
            })

    # ── FAQ Q&A — detect Q/A pairs using regex ──
    faq_section = re.search(r"Section 5.*?Q&A Format\)(.*)", text, re.DOTALL)
    if faq_section:
        faq_text = faq_section.group(1).strip()
        qa_pairs = re.findall(
            r"Q:\s*(.+?)\nA:\s*(.+?)(?=\nQ:|\Z)",
            faq_text,
            re.DOTALL
        )
        for i, (question, answer) in enumerate(qa_pairs):
            q = question.strip().replace("\n", " ")
            a = answer.strip().replace("\n", " ")
            docs.append({
                "page_content": f"Question: {q}\nAnswer: {a}",
                "metadata": {
                    "source": "pdf",
                    "page": 3,
                    "section": "faq",
                    "format": "qa_pair",
                    "doc_type": "faq",
                    "question": q[:80],
                    "faq_index": i
                }
            })

    return docs


# ── PAGE 4: Chef's Notes (unstructured text) ──────────────────────────────────
def _load_page4(page) -> list[dict]:
    docs = []
    text = page.extract_text() or ""

    notes_match = re.search(
        r"Section 6 — Chef's Notes.*?\n(.+?)(?=Section 7)",
        text, re.DOTALL
    )
    if not notes_match:
        return docs

    notes_text = notes_match.group(1).strip()

    dish_headers = [
        "Ember Ribeye Steak",
        "Spicy Dragon Noodles",
        "BBQ Pulled Pork",
        "Lava Chocolate Cake",
        "Soup of the Day"
    ]

    # Split at each dish name — they act as section headers
    pattern = "(" + "|".join(re.escape(d) for d in dish_headers) + ")"
    parts = re.split(pattern, notes_text)

    i = 1
    while i < len(parts) - 1:
        dish_name = parts[i].strip()
        note_body = parts[i + 1].strip().replace("\n", " ")
        if dish_name and note_body:
            docs.append({
                "page_content": f"Chef's note about {dish_name}:\n{note_body}",
                "metadata": {
                    "source": "pdf",
                    "page": 4,
                    "section": "chefs_notes",
                    "format": "plain_text",
                    "doc_type": "chef_note",
                    "dish_name": dish_name
                }
            })
        i += 2

    return docs


# ── PAGE 5: Nutrition inline data (pipe-separated text) ───────────────────────
def _load_page5(page) -> list[dict]:
    """
    Format on page 4-5:
        ST001 | Ember Chicken Wings
        Calories: 480 | Protein: 38g | Carbs: 12g | ...

    This is "inline structured text" — not a table, not JSON,
    but still regex-parseable. A very common real-world data format.
    """
    docs = []
    text = page.extract_text() or ""

    nutrition_blocks = re.findall(
        r"([A-Z]{2}\d{3})\s*\|\s*(.+?)\n(Calories:.+?)(?=\n[A-Z]{2}\d{3}|\Z)",
        text,
        re.DOTALL
    )

    for item_id, item_name, nutrition_line in nutrition_blocks:
        nutrition_clean = nutrition_line.strip().replace("\n", " ").replace(" | ", ", ")
        docs.append({
            "page_content": (
                f"Nutrition info for {item_name.strip()} (ID: {item_id.strip()}):\n"
                f"{nutrition_clean}"
            ),
            "metadata": {
                "source": "pdf",
                "page": 5,
                "section": "nutrition",
                "format": "inline_structured",
                "doc_type": "nutrition",
                "item_id": item_id.strip(),
                "item_name": item_name.strip()
            }
        })

    return docs


# ══════════════════════════════════════════════════════════════
#  EXCEL LOADERS — 3 sheets, 3 strategies
# ══════════════════════════════════════════════════════════════

def load_excel(excel_path: str) -> list[dict]:
    all_docs = []
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    all_docs += _load_menu_sheet(wb["Menu"])
    all_docs += _load_nutrition_sheet(wb["Nutrition"])
    all_docs += _load_hours_sheet(wb["Hours & Booking"])

    print(f"✅ Excel: extracted {len(all_docs)} documents from {excel_path}")
    return all_docs


def _load_menu_sheet(ws) -> list[dict]:
    """
    Row 1 = title (skip), Row 2 = headers, Rows 3-16 = data
    Excel has MORE columns than the PDF table: Vegan + Allergens columns
    This is why loading both files is valuable — they complement each other.
    """
    docs = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        item_id = row[0]
        if not item_id or not str(item_id).startswith(("ST", "MN", "DS", "DR")):
            continue

        name, category, price, calories, spice, veg, vegan, gf, allergens = row[1:]

        text = (
            f"Menu item: {name} (ID: {item_id})\n"
            f"Category: {category} | Price: ${price} | Calories: {calories}\n"
            f"Spice level: {spice}\n"
            f"Vegetarian: {veg} | Vegan: {vegan} | Gluten free: {gf}\n"
            f"Allergens: {allergens}"
        )
        docs.append({
            "page_content": text,
            "metadata": {
                "source": "excel_menu_sheet",
                "sheet": "Menu",
                "format": "excel_row",
                "doc_type": "menu_item",
                "item_id": str(item_id),
                "item_name": str(name),
                "category": str(category).lower(),
                "price": float(price) if price else 0,
                "is_vegetarian": str(veg),
                "is_vegan": str(vegan),
                "is_gluten_free": str(gf),
                "allergens": str(allergens)
            }
        })
    return docs


def _load_nutrition_sheet(ws) -> list[dict]:
    """Row 1 = title, Row 2 = headers, Rows 3-14 = data, Row 15+ = averages (skip)"""
    docs = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        item_id = row[0]
        if not item_id or not str(item_id).startswith(("ST", "MN", "DS", "DR")):
            continue

        item_name, calories, protein, carbs, fat, fiber, sodium, nuts = row[1:]
        text = (
            f"Nutrition facts for {item_name} (ID: {item_id}):\n"
            f"Calories: {calories} | Protein: {protein}g | Carbs: {carbs}g | "
            f"Fat: {fat}g | Fiber: {fiber}g | Sodium: {sodium}mg\n"
            f"Contains nuts: {nuts}"
        )
        docs.append({
            "page_content": text,
            "metadata": {
                "source": "excel_nutrition_sheet",
                "sheet": "Nutrition",
                "format": "excel_row",
                "doc_type": "nutrition",
                "item_id": str(item_id),
                "item_name": str(item_name),
                "calories": int(calories) if calories else 0,
                "contains_nuts": str(nuts)
            }
        })
    return docs


def _load_hours_sheet(ws) -> list[dict]:
    """Row 1 = title, Row 2 = headers, Rows 3-9 = one per day"""
    docs = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        day = row[0]
        if not day:
            continue
        status, opens, closes, notes = row[1], row[2], row[3], row[4]

        if str(status).lower() == "closed":
            text = f"Opening hours — {day}: CLOSED. Note: {notes}"
        else:
            text = f"Opening hours — {day}: Open {opens} to {closes}. Note: {notes}"

        docs.append({
            "page_content": text,
            "metadata": {
                "source": "excel_hours_sheet",
                "sheet": "Hours & Booking",
                "format": "excel_row",
                "doc_type": "hours",
                "day": str(day),
                "status": str(status)
            }
        })
    return docs


# ══════════════════════════════════════════════════════════════
#  MASTER LOADER
# ══════════════════════════════════════════════════════════════

def load_all_documents(
    pdf_path: str = "./spice_and_ember_data.pdf",
    excel_path: str = "./spice_and_ember_menu.xlsx"
) -> list[dict]:
    all_docs = []
    all_docs += load_pdf(pdf_path)
    all_docs += load_excel(excel_path)

    print(f"\n{'='*50}")
    print(f"📦 TOTAL DOCUMENTS LOADED: {len(all_docs)}")

    format_counts = {}
    for doc in all_docs:
        fmt = doc["metadata"].get("format", "unknown")
        format_counts[fmt] = format_counts.get(fmt, 0) + 1

    print("\n   Format breakdown:")
    for fmt, count in sorted(format_counts.items(), key=lambda x: -x[1]):
        print(f"   ├── {fmt}: {count} docs")
    print(f"{'='*50}\n")

    return all_docs


# ── RUN & TEST ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    docs = load_all_documents()

    print("── 1 sample from each format ──\n")
    seen = set()
    for doc in docs:
        fmt = doc["metadata"].get("format")
        if fmt not in seen:
            seen.add(fmt)
            print(f"FORMAT : {fmt}")
            print(f"SOURCE : {doc['metadata'].get('source')}")
            print(f"TEXT   :\n{doc['page_content'][:220]}")
            print(f"META   : {doc['metadata']}")
            print("-" * 50 + "\n")
